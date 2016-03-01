# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import os
import time
from datetime import date
import sys
import django

sys.path.append("/opt/web_projects/analytic_tools/")
os.environ['DJANGO_SETTINGS_MODULE'] = 'analytic_tools.settings'
django.setup()

from reporting.models import Transaction, ProductFromDB, Transport, Warehouse

def loading(fileName):
    path = ('/opt/files_store/Transactions/Input/%s.xlsx') % (fileName)
    wb = load_workbook(path, read_only = True)
    ws = wb.active
    return ws

def parsing(fileName, endDate, shift = 0):
    ws = loading(fileName)
    for row in ws['A1':'Z1048577']:
        date            = list(row)[shift + 2].value
        productCode     = list(row)[shift + 10].value
        transportCode   = list(row)[shift + 12].value
#       <--- replace master data --->
#        if transportCode == None:
#            transportCode = ''
        try:
            transportCode = str(int(transportCode))
        except:
            pass
#       <--------------------------->
        warehouseCode   = list(row)[shift + 3].value
#       <--- replace master data --->
        if warehouseCode == 1:
            warehouseCode = 'U104'
        if warehouseCode == 5:
            warehouseCode = 'U103'
#       <--------------------------->
        shippedVolume   = list(row)[shift + 14].value
        shippedAmount   = list(row)[shift + 15].value
        intercompany = [14003280]
        try:
            if date.date() < endDate and int(productCode) not in intercompany :
                dataArray.append([date.date(),
                                  int(productCode),
                                  transportCode,
                                  warehouseCode,
                                  shippedVolume,
                                  shippedAmount])
        except:
            pass

def exception(modelName, number):
    if item[number] not in modelName.objects.all().values_list('code', flat = True) and item[number] not in exceptions and item[number] != None:
        exceptions.append(item[number])    

def assignment(modelName, number):
    try:
        return modelName.objects.get(code = item[number])
    except:
        return None
    

#<--- testing function --->
def unique(column):
    valTuple = []
    for item in dataArray:
        if item[column] not in valTuple and item[column] != None:
            valTuple.append(item[column])
    return sorted(valTuple)
#<------------------------>

if __name__ == '__main__':
#   <--- rawinput --->
    year = 2016
    month = 2
    day = 26
#   <---------------->    
    clc = time.time()
    dataArray = []
    endDate = date(year, month, day)
    parsing('150914_Raw Data for Daily Report 2015', endDate)
    parsing('16.02.25', endDate, -1)
#    print(unique(<column>))
    Transaction.objects.filter(date__gte = date(year, month, 1)).delete()
    print('Transactions total count \t= %s' % len(dataArray))
    exceptions = []
    success = 0
    for item in dataArray:
        try:
            Transaction.objects.create(date          = item[0],
                                       productCode   = assignment(ProductFromDB, 1),
                                       transportCode = assignment(Transport, 2),
                                       warehouseCode = assignment(Warehouse, 3),
                                       shippedVolume = item[4],
                                       shippedAmount = item[5])
            success += 1
        except:
            exception(ProductFromDB, 1)
            exception(Transport, 2)
            exception(Warehouse, 3)
    for item in exceptions:
        print('Object %s not found' % item)
    print('Transactions success count \t= %s' % success)
    clc = time.time() - clc
    print('Script working time \t\t= {0:.2f} seconds'.format(clc))