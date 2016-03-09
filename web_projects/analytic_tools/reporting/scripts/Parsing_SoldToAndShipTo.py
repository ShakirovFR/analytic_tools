# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import os
import sys
import django

sys.path.append("/opt/web_projects/analytic_tools/")
os.environ['DJANGO_SETTINGS_MODULE'] = 'analytic_tools.settings'
django.setup()

from reporting.models import SoldToFromDB, ShipToFromDB, SoldTo, ShipTo
from django.db import IntegrityError

def extraction(ws, start, end):
    dataArray = []
    for row in ws[start:end]:
        dataArray.append([item.value for item in list(row)])
    return dataArray

def clientCard(data, client):
    for item in data:
        try:
            obj, created = client.objects.get_or_create(code = 'U' + str(item.code),
                                                        defaults = {'name':item.name})
            obj.codeFromDB.add(item)
        except IntegrityError:
            obj = client.objects.get(name = item.name)
            obj.codeFromDB.add(item)

def checkClient(model, item):
    check = model.objects.get(code = item)
    print(check.codeFromDB.values_list('code', flat = True))

def main():
    fileName = 'SoldToAndShipTo'
    path = ('/opt/files_store/MasterData/%s.xlsx') % (fileName)
    wb = load_workbook(path, read_only = True)
    
    ws1 = wb['soldTo']
    soldTo = extraction(ws1, 'A2', 'B1368')
    for item in soldTo:
        SoldToFromDB.objects.get_or_create(code     = item[0],
                                           defaults = {'name':item[1]})

    soldToFromDB = SoldToFromDB.objects.all()
    clientCard(soldToFromDB, SoldTo)

    ws2 = wb['shipTo']
    shipTo = extraction(ws2, 'A2', 'B2342')
    for item in shipTo:
        ShipToFromDB.objects.get_or_create(code     = item[0],
                                           defaults = {'name':item[1]})

    shipToFromDB = ShipToFromDB.objects.all()
    clientCard(shipToFromDB, ShipTo)

#    checkClient(ShipTo, 'U100457680')

if __name__ == '__main__':
    main()