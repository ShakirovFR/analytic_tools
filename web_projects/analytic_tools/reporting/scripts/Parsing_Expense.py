# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import os
import datetime

path = []
path.append('/opt/files_store/EBITDA/ProductionExpense/')
path.append('/opt/files_store/EBITDA/OtherExpense/')

filesList =[]
for item in path: 
    filesListNew = [list(item) for item in os.walk(item)]
    filesListNew[0][0] = item
    filesListNew[0].pop(1)
    filesList += filesListNew

def getData(directory, fileName):
    modelName = fileName.rpartition('.')[0]
    path = directory + fileName
    fullPath = os.path.expanduser(path)
    wb = load_workbook(fullPath)
    ws = wb.active
    # Максимальное кол-во строк в Excel2010 = 1048576
    for i in range(1, 1048577):
        row = ws.cell(row = i, column = 1).value
        if row == None or row == '':
            rowEnd = i
            break
    # Максимальное кол-во столбцов в Excel2010 = 16384
    columnHeader = []     
    for i in range(1, 16385):
        column = ws.cell(row = 1, column = i).value
        columnHeader.append(column)
        if column == None or column == '':
            columnEnd = i
            break
    columnHeader.pop()
    value = []
    for i in range(2, rowEnd):
        rowValues = []        
        for j in range(1, columnEnd):
            if type(ws.cell(row = i, column = j).value) == datetime.datetime:
                rowValues.append(str(ws.cell(row = i, column = j).value.date()))
            else:
                rowValues.append(ws.cell(row = i, column = j).value)
        for item in rowValues:
            if item == '' or item == None:
                index = rowValues.index(item)
                if item == '':
                    rowValues.remove('')
                else:
                    rowValues.remove(None)
                rowValues.insert(index, 'Null')
        value.append(rowValues)
    return [modelName, columnHeader, value]

allData = []
for directory in filesList:
    for fileName in directory[1]:
        allData.append(getData(directory[0], fileName))