# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import os
from datetime import datetime
from dateutil.relativedelta import relativedelta

today = datetime.now()
previousDate = (today + relativedelta(months = -1))
searchMonth = previousDate.strftime('%B')

path = '/opt/files_store/EBITDA/DistributionReport/Distribution Report AY16.xlsx'
wb = load_workbook(path, data_only = True, read_only = True)

#   Distribution Shurovo
#U101M7501	Packing         Grey
#U101M7502	Packing         White
#U101M7503	Shipping        Grey
#U101M7504	Shipping        White
#U101M7601	Logistic Road   Grey    Bulk
#U101M7602	Logistic Road   Grey    Bag
#U101M7603	Logistic Road   White
#U101M7604	Dispatch
#U101M7701	Logistic Rail   Grey    Bulk
#U101M7702	Logistic Rail   White

#   Distribution Volsk
#U102M7501	Packing
#U102M7502	Shipping
#U102M7601	Logistic Road   Grey    Bulk
#U102M7602	Logistic Road   Grey    Bag
#U102M7701	Logistic Rail   Grey    Bulk
#U102M7702	Logistic Rail   Grey    Bag
#U102M7703	Logistic Rail   API
#U102M7705	Dispatch
#U102M7801	Logistic Water

#	Distribution Ferzikovo
#U103M7501	Packing
#U103M7502	Shipping
#U103M7601	Logistic Road   Grey    Bulk
#U103M7602	Logistic Road   Grey    Bag

#	Distribution Voskresensk
#U104M7501	Packing
#U104M7502	Shipping
#U104M7601	Logistic Road   Grey    Bulk
#U104M7602	Logistic Road   Grey    Bag

#   Terminal Tatarstan
#U132M7501	Shipping        
#U132M7601	Logistic Road

#   Terminal Kazakhstan
#!!! Terminal KAZ !!!
#U140M7501	Shipping        
#U140M7601	Logistic Road

#   Distribution Shurovo
U101M7501 = wb['U101M7501']
U101M7502 = wb['U101M7502']
U101M7503 = wb['U101M7503']
U101M7504 = wb['U101M7504']
U101M7601 = wb['U101M7601']
U101M7602 = wb['U101M7602']
U101M7603 = wb['U101M7603']
U101M7604 = wb['U101M7604']
U101M7701 = wb['U101M7701']
U101M7702 = wb['U101M7702']
#   Distribution Volsk
U102M7501 = wb['U102M7501']
U102M7502 = wb['U102M7502']
U102M7601 = wb['U102M7601']
U102M7602 = wb['U102M7602']
U102M7701 = wb['U102M7701']
U102M7702 = wb['U102M7702']
U102M7703 = wb['U102M7703']
U102M7705 = wb['U102M7705']
U102M7801 = wb['U102M7801']
#	Distribution Ferzikovo
U103M7501 = wb['U103M7501']
U103M7502 = wb['U103M7502']
U103M7601 = wb['U103M7601']
U103M7602 = wb['U103M7602']
#	Distribution Voskresensk
U104M7501 = wb['U104M7501']
U104M7502 = wb['U104M7502']
U104M7601 = wb['U104M7601']
U104M7602 = wb['U104M7602']
#   Terminal Tatarstan
U132M7501 = wb['U132M7501']
U132M7601 = wb['U132M7601']
#   Terminal Kazakhstan
Terminal_KAZ = wb['Terminal KAZ']

#   Total distribution
TOTAL = wb['TOTAL']
#   Corporate logistic
CL_HM = wb['CL HM']
#   Clinker
#U101M7703  Shurovo
U101M7703 = wb['U101M7703']
#U102M7704  Volsk
U102M7704 = wb['U102M7704']
#U103M7603  Ferzikovo
U103M7603 = wb['U103M7603']
#U104M7603  Voskresensk
U104M7603 = wb['U104M7603']

# searchMonth = 'January'

for i in range(1, 16385):
    if TOTAL.cell(row = 1, column = i).value == searchMonth:
        month = i + 1
        break

#for j in range(1, 1048577):
#    if TOTAL.cell(row = j, column = 3).value == '<SearchItem>':
#        print(j)
#        break

#              Actual Cost
all = 275
#   3225500	Outbound freight 
dom = 81
#   3225510	LG-Outbound freight export 
exp = 83
#   3225503	LG-Outbound freight - rentals rail 
rail1 = 82
#   3225550	Subcontracting costs for Trucks or rail cars 
rail2 = 85

#   All                 : row = 'Actual Cost', column = 'January' + 1 ('I275')
#   Freight domestic    : row = '3225500', column = 'January' + 1 ('I81')
#   Freight export      : row = '3225510', column = 'January' + 1 ('I83')
#   Wagons rent         : row = '3225503' + '3225550', column = 'January' + 1 ('I82' + 'I85')
#   All other           : All - (Freight domestic + Freight export + Wagons rent)

def val(sheet, row):
    if sheet.cell(row = row, column = month).value == '' or sheet.cell(row = row, column = month).value == None:
        value = 0.0
    else:
        value = sheet.cell(row = row, column = month).value
    return value

checksum = (val(TOTAL, all) - val(CL_HM, all) - val(U101M7703, all) - val(U102M7704, all) - val(U103M7603, all) - val(U104M7603, all)) * 1000

valuesList = []
#U101M7501	All
valuesList.append(val(U101M7501, all))
#U101M7502	All
valuesList.append(val(U101M7502, all))
#U101M7503	All
valuesList.append(val(U101M7503, all))
#U101M7504	All
valuesList.append(val(U101M7504, all))
#U101M7601	All
valuesList.append(val(U101M7601, all))
#U101M7602	All
valuesList.append(val(U101M7602, all))
#U101M7603	All
valuesList.append(val(U101M7603, all))
#U101M7604	All
valuesList.append(val(U101M7604, all))
#U101M7701	Freight domestic
valuesList.append(val(U101M7701, dom))
#U101M7701	Wagons rent
valuesList.append(val(U101M7701, rail1) + val(U101M7701, rail2))
#U101M7701	All other
valuesList.append(val(U101M7701, all) - val(U101M7701, dom) - val(U101M7701, rail1) - val(U101M7701, exp) - val(U101M7701, rail2))
#U101M7702	Freight domestic
valuesList.append(val(U101M7702, dom))
#U101M7702	Freight export
valuesList.append(val(U101M7702, exp))
#U101M7702	Wagons rent
valuesList.append(val(U101M7702, rail1) + val(U101M7702, rail2))
#U101M7702	All other
valuesList.append(val(U101M7702, all) - val(U101M7702, dom) - val(U101M7702, rail1) - val(U101M7702, exp) - val(U101M7702, rail2))
#U102M7501	All
valuesList.append(val(U102M7501, all))
#U102M7502	All
valuesList.append(val(U102M7502, all))
#U102M7601	All
valuesList.append(val(U102M7601, all))
#U102M7602	All
valuesList.append(val(U102M7602, all))
#U102M7701	Freight domestic
valuesList.append(val(U102M7701, dom))
#U102M7701	Freight export
valuesList.append(val(U102M7701, exp))
#U102M7701	Wagons rent
valuesList.append(val(U102M7701, rail1) + val(U102M7701, rail2))
#U102M7701	All other
valuesList.append(val(U102M7701, all) - val(U102M7701, dom) - val(U102M7701, rail1) - val(U102M7701, exp) - val(U102M7701, rail2))
#U102M7702	Freight domestic
valuesList.append(val(U102M7702, dom))
#U102M7702	Freight export
valuesList.append(val(U102M7702, exp))
#U102M7702	Wagons rent
valuesList.append(val(U102M7702, rail1) + val(U102M7702, rail2))
#U102M7702	All other
valuesList.append(val(U102M7702, all) - val(U102M7702, dom) - val(U102M7702, rail1) - val(U102M7702, exp) - val(U102M7702, rail2))
#U102M7703	Freight domestic
valuesList.append(val(U102M7703, dom))
#U102M7703	Freight export
valuesList.append(val(U102M7703, exp))
#U102M7703	Wagons rent
valuesList.append(val(U102M7703, rail1) + val(U102M7703, rail2))
#U102M7703	All other
valuesList.append(val(U102M7703, all) - val(U102M7703, dom) - val(U102M7703, rail1) - val(U102M7703, exp) - val(U102M7703, rail2))
#U102M7705	All
valuesList.append(val(U102M7705, all))
#U102M7801	All
valuesList.append(val(U102M7801, all))
#U103M7501	All
valuesList.append(val(U103M7501, all))
#U103M7502	All
valuesList.append(val(U103M7502, all))
#U103M7601	All
valuesList.append(val(U103M7601, all))
#U103M7602	All
valuesList.append(val(U103M7602, all))
#U104M7501	All
valuesList.append(val(U104M7501, all))
#U104M7502	All
valuesList.append(val(U104M7502, all))
#U104M7601	All
valuesList.append(val(U104M7601, all))
#U104M7602	All
valuesList.append(val(U104M7602, all))
#U132M7501	All
valuesList.append(val(U132M7501, all))
#U132M7601	All
valuesList.append(val(U132M7601, all))
#Terminal_KAZ All
valuesList.append(val(Terminal_KAZ, all))
valuesList.append(0.0)

totalsum = 0
for i in range(len(valuesList)):
    valuesList[i] *= 1000
    totalsum += valuesList[i]

# print(valuesList)
# print('Difference =', totalsum - checksum)

#wb = load_workbook(os.path.expanduser('~/Desktop/EBITDA/Distribution.xlsx'))
#ws = wb.active
#for i in range(len(valuesList)):
#    ws.cell(row = i + 2, column = 8).value = valuesList[i]
#wb.save(os.path.expanduser('~/Desktop/EBITDA/Distribution.xlsx'))